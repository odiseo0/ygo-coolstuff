from dataclasses import dataclass
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook

from src.models import cards
from src.models.db_models import CollectionItem as DbCollectionItem
from src.utils.constants import BASE_URL, EXCEL_TEMPLATE_FILENAME, EXPORT_PLATFORM_NAME
from src.utils.utils import trim_card_name


def get_default_template_path() -> Path:
    """Template path from project root."""
    project_root = Path(__file__).resolve().parent.parent.parent
    return project_root / EXCEL_TEMPLATE_FILENAME


EXCEL_HEADER_CANTIDAD = "Cantidad"
EXCEL_HEADER_NOMBRE = "Nombre de la Carta"
EXCEL_HEADER_CODIGO = "Codigo"
EXCEL_HEADER_RAREZA = "Rareza"
EXCEL_HEADER_PLATAFORMA = "Plataforma"
EXCEL_HEADER_ENLACE = "Enlace de la Carta"
EXCEL_HEADER_EXTRA = "Extra"

EXPECTED_HEADERS = (
    EXCEL_HEADER_CANTIDAD,
    EXCEL_HEADER_NOMBRE,
    EXCEL_HEADER_CODIGO,
    EXCEL_HEADER_RAREZA,
    EXCEL_HEADER_PLATAFORMA,
    EXCEL_HEADER_ENLACE,
    EXCEL_HEADER_EXTRA,
)

DATA_START_ROW = 2


@dataclass
class ExportRow:
    quantity: int
    card_name: str
    code: str
    rarity: str


def build_card_link(card_name: str) -> str:
    trimmed = trim_card_name(card_name).strip()

    if not trimmed:
        return ""
    encoded = quote(trimmed, safe="").replace("%20", "+")

    return f"{BASE_URL}{encoded}"


def cards_item_to_export_rows(
    items: list[cards.CollectionItem],
) -> list[ExportRow]:
    return [
        ExportRow(
            quantity=item.qty,
            card_name=item.name,
            code=item.code,
            rarity=item.rarity,
        )
        for item in items
    ]


def db_items_to_export_rows(items: list[DbCollectionItem]) -> list[ExportRow]:
    return [
        ExportRow(
            quantity=item.card_quantity,
            card_name=item.card_name,
            code=item.card_code,
            rarity=item.card_rarity,
        )
        for item in items
    ]


def _column_index_by_header(sheet, header: str) -> int | None:
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value is not None and str(cell.value).strip() == header:
            return col_idx
    return None


def export_collection_to_excel(
    rows: list[ExportRow],
    template_path: Path,
    output_path: Path,
) -> None:
    if not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    workbook = load_workbook(filename=template_path, read_only=False)
    sheet = workbook.active
    if sheet is None:
        workbook.close()
        raise ValueError("Template has no active sheet")

    indices: dict[str, int] = {}
    for header in EXPECTED_HEADERS:
        idx = _column_index_by_header(sheet, header)
        if idx is None:
            workbook.close()
            raise ValueError(f"Template missing header: {header}")
        indices[header] = idx

    for row_offset, export_row in enumerate(rows):
        excel_row = DATA_START_ROW + row_offset
        sheet.cell(row=excel_row, column=indices[EXCEL_HEADER_CANTIDAD]).value = (
            export_row.quantity
        )
        sheet.cell(row=excel_row, column=indices[EXCEL_HEADER_NOMBRE]).value = (
            trim_card_name(export_row.card_name)
        )
        sheet.cell(row=excel_row, column=indices[EXCEL_HEADER_CODIGO]).value = (
            export_row.code
        )
        sheet.cell(row=excel_row, column=indices[EXCEL_HEADER_RAREZA]).value = (
            export_row.rarity
        )
        sheet.cell(row=excel_row, column=indices[EXCEL_HEADER_PLATAFORMA]).value = (
            EXPORT_PLATFORM_NAME
        )
        sheet.cell(row=excel_row, column=indices[EXCEL_HEADER_ENLACE]).value = (
            build_card_link(export_row.card_name)
        )
        sheet.cell(row=excel_row, column=indices[EXCEL_HEADER_EXTRA]).value = ""

    workbook.save(filename=output_path)
    workbook.close()
